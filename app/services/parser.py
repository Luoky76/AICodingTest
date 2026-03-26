"""Content parsing service.

Handles parsing of email and Excel inputs into raw text.
"""

from __future__ import annotations

import base64
import io

from app.models.schemas import OrderInput, ParsedContent


def parse_content(order_input: OrderInput) -> ParsedContent:
    """Parse input content into raw text.

    Args:
        order_input: The order input (email or excel type).

    Returns:
        ParsedContent with extracted raw text.

    Raises:
        ValueError: If input type is unsupported or content cannot be parsed.
    """
    if order_input.type == "email":
        return _parse_email(order_input)
    elif order_input.type == "excel":
        return _parse_excel(order_input)
    else:
        raise ValueError(f"Unsupported input type: {order_input.type}")


def _parse_email(order_input: OrderInput) -> ParsedContent:
    """Extract text from email input."""
    parts: list[str] = []

    if order_input.subject:
        parts.append(f"Subject: {order_input.subject}")

    if order_input.body:
        parts.append(f"Body:\n{order_input.body}")

    if order_input.attachments:
        for attachment in order_input.attachments:
            parts.append(f"Attachment ({attachment.fileName}):")
            content = attachment.fileContent

            # Try to parse Excel attachments
            if attachment.fileName.endswith((".xlsx", ".xls")):
                try:
                    decoded = base64.b64decode(content)
                    text = _read_excel_bytes(decoded)
                    parts.append(text)
                except Exception:
                    parts.append(content)
            else:
                # Treat as text content
                try:
                    decoded_text = base64.b64decode(content).decode("utf-8")
                    parts.append(decoded_text)
                except Exception:
                    parts.append(content)

    raw_text = "\n\n".join(parts)
    return ParsedContent(rawText=raw_text)


def _parse_excel(order_input: OrderInput) -> ParsedContent:
    """Extract text from Excel input."""
    if not order_input.fileContent:
        raise ValueError("Excel input must have fileContent")

    try:
        decoded = base64.b64decode(order_input.fileContent)
    except Exception as e:
        raise ValueError(f"Failed to decode Excel content: {e}")

    text = _read_excel_bytes(decoded)
    return ParsedContent(rawText=text)


def _read_excel_bytes(data: bytes) -> str:
    """Read Excel bytes and convert to text representation."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required to parse Excel files")

    wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    parts: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"Sheet: {sheet_name}")

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Use first row as headers if available
        headers = [str(cell) if cell is not None else "" for cell in rows[0]]
        parts.append(" | ".join(headers))

        for row in rows[1:]:
            values = [str(cell) if cell is not None else "" for cell in row]
            parts.append(" | ".join(values))

    wb.close()
    return "\n".join(parts)
